# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.error('openpyxl library is not installed')
    raise UserError('openpyxl library is required for import functionality')


class ImportCommunicationCodesWizard(models.TransientModel):
    _name = 'import.communication.codes.wizard'
    _description = 'استيراد شفرات الاتصال من Excel'
    _rec_name = 'filename'

    file = fields.Binary(
        string='ملف Excel',
        required=True,
        attachment=False,
    )

    filename = fields.Char(
        string='اسم الملف',
    )

    import_mode = fields.Selection(
        selection=[
            ('create', 'إنشاء سجلات جديدة'),
            ('update', 'تحديث السجلات الموجودة'),
        ],
        string='وضع الاستيراد',
        default='create',
        required=True,
    )

    error_lines = fields.Text(
        string='سجل الأخطاء',
        readonly=True,
    )

    success_count = fields.Integer(
        string='السجلات الناجحة',
        readonly=True,
        default=0,
    )

    error_count = fields.Integer(
        string='السجلات الفاشلة',
        readonly=True,
        default=0,
    )

    state = fields.Selection(
        selection=[
            ('draft', 'مسودة'),
            ('done', 'تم'),
        ],
        string='الحالة',
        default='draft',
        readonly=True,
    )

    CODE_SYSTEM_MAP = {
        'prepaid': 'prepaid',
        'monthly_invoice': 'monthly_invoice',
        'other': 'other',
        'Prepaid': 'prepaid',
        'Monthly Invoice': 'monthly_invoice',
        'Other': 'other',
    }

    CODE_STATUS_MAP = {
        'in_stock': 'in_stock',
        'delivered': 'delivered',
        'suspended': 'suspended',
        'cancelled': 'cancelled',
        'In Stock': 'in_stock',
        'Delivered': 'delivered',
        'Suspended': 'suspended',
        'Cancelled': 'cancelled',
    }

    CODE_VERSION_MAP = {
        'original': 'original',
        'new_version': 'new_version',
        'Original': 'original',
        'New Version': 'new_version',
    }

    REQUIRED_COLUMNS = [
        'Employee Name',
        'Code Number',
    ]

    def action_import(self):
        self.ensure_one()

        if not self.file:
            raise UserError('Please select an Excel file.')

        if not self.filename or not self.filename.endswith('.xlsx'):
            raise UserError('Please select a valid Excel file (.xlsx).')

        try:
            errors, success_count, error_count = self._process_file()

            error_message = '\n'.join(errors) if errors else 'Import completed successfully!'

            self.write({
                'error_lines': error_message,
                'success_count': success_count,
                'error_count': error_count,
                'state': 'done',
            })

            return self._reload_wizard()

        except Exception as e:
            _logger.exception('Error during import')
            raise UserError(f'Error reading file: {str(e)}')

    def _process_file(self):
        errors = []
        success_count = 0
        error_count = 0

        file_content = base64.b64decode(self.file)
        workbook = self._load_workbook(file_content)
        sheet = workbook.active

        validation_errors = self._validate_headers(sheet)
        if validation_errors:
            errors.extend(validation_errors)
            return errors, 0, len(validation_errors)

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] and not row[5]:
                continue

            try:
                result = self._process_row(row, row_idx)
                if result.get('error'):
                    errors.append(result['error'])
                    error_count += 1
                else:
                    success_count += 1

            except Exception as e:
                errors.append(f'Row {row_idx}: Data error - {str(e)}')
                error_count += 1

        return errors, success_count, error_count

    def _load_workbook(self, file_content):
        return load_workbook(io.BytesIO(file_content))

    def _validate_headers(self, sheet):
        errors = []
        headers = [cell.value for cell in sheet[1]]

        for col in self.REQUIRED_COLUMNS:
            if col not in headers:
                errors.append(f'Column "{col}" not found in file.')

        return errors

    def _process_row(self, row, row_idx):
        employee_name = row[0] if row[0] else ''
        code_number = str(row[5]).strip() if row[5] else ''

        if not code_number:
            return {'error': f'Row {row_idx}: Code number is required.'}

        employee_id = self._find_or_create_employee(employee_name, row_idx)
        if not employee_id:
            return {'error': f'Row {row_idx}: Employee "{employee_name}" not found.'}

        code_system = self._map_value(row[6], self.CODE_SYSTEM_MAP, 'system', row_idx)
        code_status = self._map_value(row[8], self.CODE_STATUS_MAP, 'status', row_idx)
        code_version = self._map_value(row[9], self.CODE_VERSION_MAP, 'version', row_idx)

        vals = {
            'employee_id': employee_id,
            'city': row[4] if row[4] else '',
            'code_number': code_number,
            'code_system': code_system,
            'monthly_balance': self._parse_float(row[7]),
            'code_status': code_status,
            'code_version': code_version,
        }

        existing_code = self._find_existing_code(code_number)
        if existing_code:
            if self.import_mode == 'create':
                return {'error': f'Row {row_idx}: Code "{code_number}" already exists.'}
            existing_code.write(vals)
        else:
            if self.import_mode == 'update':
                return {'error': f'Row {row_idx}: Code "{code_number}" not found for update.'}
            self.env['communication.codes'].create(vals)

        return {'success': True}

    def _find_or_create_employee(self, employee_name, row_idx):
        if not employee_name:
            return False

        employee = self.env['hr.employee'].search([
            ('name', 'ilike', employee_name)
        ], limit=1)

        if employee:
            return employee.id

        _logger.warning(f'Employee "{employee_name}" not found at row {row_idx}')
        return False

    def _find_existing_code(self, code_number):
        return self.env['communication.codes'].search([
            ('code_number', '=', code_number)
        ], limit=1)

    def _map_value(self, value, mapping, field_type, row_idx):
        if not value:
            defaults = {
                'system': 'prepaid',
                'status': 'in_stock',
                'version': 'original',
            }
            return defaults.get(field_type, '')

        mapped_value = mapping.get(value)
        if not mapped_value:
            _logger.warning(
                f'Row {row_idx}: Invalid {field_type} value "{value}"'
            )
            defaults = {
                'system': 'prepaid',
                'status': 'in_stock',
                'version': 'original',
            }
            return defaults.get(field_type, '')

        return mapped_value

    def _parse_float(self, value):
        if not value:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _reload_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'import.communication.codes.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_reset(self):
        self.write({
            'state': 'draft',
            'file': False,
            'filename': False,
            'error_lines': False,
            'success_count': 0,
            'error_count': 0,
        })
        return self._reload_wizard()
