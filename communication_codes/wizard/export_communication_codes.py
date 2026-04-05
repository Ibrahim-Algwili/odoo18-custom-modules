# -*- coding: utf-8 -*-

import base64
import io
import logging
import re
from datetime import datetime

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.error("openpyxl library is not installed")
    raise UserError("openpyxl library is required for export functionality")


class ExportCommunicationCodesWizard(models.TransientModel):
    _name = "export.communication.codes.wizard"
    _description = "تصدير شفرات الاتصال إلى Excel"
    _rec_name = "filename"

    CODE_LIMIT = 10000

    code_ids = fields.Many2many(
        "communication.codes",
        string="الشفرات",
        default=lambda self: self._default_code_ids(),
    )

    export_all = fields.Boolean(
        string="تصدير جميع السجلات",
        default=True,
    )

    file = fields.Binary(
        string="ملف Excel",
        readonly=True,
    )

    filename = fields.Char(
        string="اسم الملف",
        readonly=True,
    )

    state = fields.Selection(
        selection=[
            ("draft", "مسودة"),
            ("done", "تم"),
        ],
        string="الحالة",
        default="draft",
        readonly=True,
    )

    CODE_SYSTEM_LABELS = {
        "prepaid": "Prepaid",
        "monthly_invoice": "Monthly Invoice",
        "other": "Other",
    }

    CODE_STATUS_LABELS = {
        "in_stock": "In Stock",
        "delivered": "Delivered",
        "suspended": "Suspended",
        "cancelled": "Cancelled",
    }

    CODE_VERSION_LABELS = {
        "original": "Original",
        "new_version": "New Version",
    }

    COLUMNS = [
        "Sequence Number",
        "Employee Name",
        "Job Position",
        "Company",
        "Branch",
        "City",
        "Code Number",
        "Code System",
        "Monthly Balance",
        "Code Status",
        "Code Version",
        "Delivery Date",
        "Notes",
    ]

    COLUMN_WIDTHS = [15, 20, 20, 20, 20, 15, 20, 20, 15, 15, 20, 20, 30]

    def _default_code_ids(self):
        active_ids = self._context.get("active_ids", [])
        return active_ids

    def action_export(self):
        self.ensure_one()

        codes = self._get_codes()
        if not codes:
            raise UserError("No data to export.")

        if len(codes) > self.CODE_LIMIT:
            raise UserError(
                f"Export limit exceeded. Maximum {self.CODE_LIMIT} records allowed."
            )

        try:
            file_data, filename = self._generate_excel(codes)

            self.write(
                {
                    "file": file_data,
                    "filename": filename,
                    "state": "done",
                }
            )

            return self._reload_wizard()

        except Exception as e:
            _logger.exception("Error during export")
            raise UserError(f"Error generating file: {str(e)}")

    def _get_codes(self):
        if self.export_all:
            return self.env["communication.codes"].search([])
        return self.code_ids

    def _generate_excel(self, codes):
        wb = Workbook()
        ws = wb.active
        ws.title = "Communication Codes"

        self._apply_header_styles(ws)
        self._write_headers(ws)
        self._write_data(ws, codes)
        self._set_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_data = base64.b64encode(output.getvalue())
        filename = (
            f'Communication_Codes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

        return file_data, filename

    def _apply_header_styles(self, ws):
        ws.header_font = Font(bold=True, color="FFFFFF")
        ws.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws.header_alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )
        ws.data_alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )

    def _write_headers(self, ws):
        for col_idx, header in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = ws.header_font
            cell.fill = ws.header_fill
            cell.alignment = ws.header_alignment
        ws.row_dimensions[1].height = 30

    def _write_data(self, ws, codes):
        for row_idx, code in enumerate(codes, start=2):
            row_values = [
                code.name or "",
                code.employee_id.name or "",
                code.job_id.name or "",
                code.company_id.name or "",
                code.branch_id.name or "",
                code.city or "",
                code.code_number or "",
                self.CODE_SYSTEM_LABELS.get(code.code_system, ""),
                code.monthly_balance or 0,
                self.CODE_STATUS_LABELS.get(code.code_status, ""),
                self.CODE_VERSION_LABELS.get(code.code_version, ""),
                str(code.delivery_date) if code.delivery_date else "",
                self._sanitize_html(code.notes),
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = ws.data_alignment

            ws.row_dimensions[row_idx].height = 25

    def _sanitize_html(self, html_content):
        if not html_content:
            return ""
        text = re.sub(r"<[^>]+>", " ", html_content)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _set_column_widths(self, ws):
        for col_idx, width in enumerate(self.COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def action_download(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{self._name}/{self.id}/file/{self.filename}?download=true",
            "target": "self",
        }

    def _reload_wizard(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": "export.communication.codes.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_reset(self):
        self.write(
            {
                "state": "draft",
                "file": False,
                "filename": False,
            }
        )
        return self._reload_wizard()
